# -*- coding:utf-8 -*-
from openpyxl import load_workbook
from openpyxl import Workbook
from tkinter import *
import json
import time
import os
import urllib3
import hashlib

provinces = None	#暂时没有使用
citys = None
orgs = None

mubans = {
	1:'func_a',		#慈铭模板
	2:'func_b',		#第一健康模板
	3:'func_c',		#公立医院模板
	4:'func_d',		#南山医院模板
	5:'func_e',		#港大医院模板
	6:'func_f',		#儿童医院模板
	7:'func_g',		#妇幼保健院模板
	8:'func_h',		#乐检查模板
	9:'func_i',		#成都市第三人民医院
	10:'func_j',	#爱康国宾模板
	11:'func_k'		#爱康国宾模板
}

org_dict = {}	#当前选中的机构信息
json_str = ''	#预约信息

ca_height = 0	#canvas中的内容高度

class MyFrame():
	def __init__(self):
		self.create_frame()

	def create_frame(self):
		self.f = Tk()
		self.f.iconbitmap('./param/tj.ico')
		Label(self.f,text='体检预约').pack()
		self.f.minsize(900,700)
		self.f.maxsize(900,700)
		self.f.title("获取预约信息")
		self.v = IntVar()
		self.v.set(0)
		
		self.sc = Scrollbar(self.f)
		#self.sc.set(0.5,1)
		self.sc.pack(fill=Y,side=RIGHT)
		
		self.l_ordersn = Label(self.f,text="订单号:")
		self.l_ordersn.pack()
		self.t_ordersn = Entry(self.f,width=20)
		self.t_ordersn.pack()
		
		self.ca = Canvas(self.f,yscrollcommand=self.sc.set,relief='flat')
		self.frame = Frame(self.ca,borderwidth=1,bg='yellow')
		
		self.addlabelframe();
		self.frame.pack()
		self.ca.create_window(450,950,window=self.frame)
		
		self.sc.config(command=self.ca.yview)
		self.ca.config(scrollregion=(0,0,600,2000))
		self.ca.pack(fill='both',expand=True)
		
		self.b_lf = LabelFrame(self.f,text='',borderwidth=1,highlightbackground='red')
		self.b_lf.pack(fill="x",expand='yes')
		
		self.b_get = Button(self.b_lf,bg='pink',text="获取",command= lambda :b_get_info(self)) #按钮
		self.b_get.grid(row=0,column=0,pady=20,padx=150,)
	
		self.b_exp = Button(self.b_lf,bg='pink',text='导出',state='disabled',command=lambda :b_ex_info(self)) #导出
		self.b_exp.grid(row=0,column=1,pady=20,padx=150,)
		
		self.l_get_tip = Label(self.b_lf,fg='green',text="")
		self.l_get_tip.grid(row=1,column=0,)
		
		self.l_ex_tip = Label(self.b_lf,fg='green',text='')
		self.l_ex_tip.grid(row=1,column=1)
		
		self.f.mainloop()
		
	def addlabelframe(self):
		global ca_height
		for city in citys:	
			
			lframe = LabelFrame(self.frame,fg='SlateBlue',text=city.get('cname'))
			self.addelem(lframe,city.get('id'))
			lframe.pack(fill='both',expand=True)

	def addelem(self,lf,cid):
		global ca_height
		rline = 0
		cline = 0
		for org in orgs:
			if(org.get('cid') == cid):
				#print(org.get('id'))
				rabutton = Radiobutton(lf,text=org.get('oname'),variable=self.v,value=org.get('id'),command=lambda x=org.get('id') :rb_click(self,x))
				#rabutton.pack(fill='both',expand='yes')
				if(cline / 4 == 1):
					rline += 1
					cline = 0
				rabutton.grid(row=rline,column=cline,sticky=W)
				cline += 1
	
def rb_click(f,oid):
	global org_dict
	for org in orgs:
		if org.get('id') == oid:
			org_dict = org
	print(org_dict)
	f.l_get_tip['text'] = ''
	f.l_ex_tip['text'] = ''
	f.b_exp.configure(state='normal')

def b_get_info(f):
	global json_str
	print(f.t_ordersn.get())
	json_str = get_appoint_info(f.t_ordersn.get())
	print(json_str)
	f.l_ex_tip['text'] = ''
	if(json_str and json_str != 'error' and json_str != '[]'):
		f.l_get_tip['text'] = '获取成功'
		#f.b_exp.configure(state='normal')
	else:
		f.l_get_tip['fg'] = 'red'
		f.l_get_tip['text'] = '获取失败'
	
def b_ex_info(f):
	global json_str,org_dict
	f.l_get_tip['text'] = ''
	export_appoint_info(json_str,org_dict)
	f.l_ex_tip['text'] = '导出成功'
	f.b_exp.configure(state='disabled')
	
def get_appoint_info(order_sn):
	re = None
	code = '1389jxxxxxxx0529'		#
	hl = hashlib.md5()
	hl.update(code.encode(encoding='utf-8'))
	md5code = hl.hexdigest()
	
	url = 'https://=%s&code=%s' % (order_sn,md5code)

	#url='www.baidu.com'
	http = urllib3.PoolManager()
	#user_agent = 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Mobile Safari/537.36'
	try:
		re = http.request('get',
				url,
				retries=False	#不使用重定向
			)
		json_str = re.data.decode()		# bytes 转 string
		return json_str
	except Exception as err:
		print("获取预约信息出错..")
		print(err)
		return ''
	finally:
		if re:
			re.release_conn()

#慈铭模板
def func_a(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('time')
				t_sheet['B%d' % i].value = dict_info.get('name')
				t_sheet['C%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['D%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['E%d' % i].value = dict_info.get('manid')
				t_sheet['F%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['G%d' % i].value = dict_info.get('goods_name')
				t_sheet['H%d' % i].value = dict_info.get('phone')
				t_sheet['I%d' % i].value = dict_info.get('org_name')
				break
	t_wb.save(filename)

#第一健康模板
def func_b(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = dict_info.get('manid')
				t_sheet['D%d' % i].value = dict_info.get('phone')
				t_sheet['E%d' % i].value = dict_info.get('goods_name')
				t_sheet['F%d' % i].value = dict_info.get('org_name')
				t_sheet['G%d' % i].value = dict_info.get('time')
				t_sheet['H%d' % i].value = ''
				break
	t_wb.save(filename)

#公立医院模板	
def func_c(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['D%d' % i].value = dict_info.get('goods_name')
				t_sheet['E%d' % i].value = dict_info.get('phone')
				t_sheet['F%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['G%d' % i].value = dict_info.get('manid')
				#t_sheet['H%d' % i].value = ''
				t_sheet['H%d' % i].value = dict_info.get('shop_price')
				t_sheet['I%d' % i].value = dict_info.get('org_name')
				t_sheet['J%d' % i].value = dict_info.get('time')
				break
	t_wb.save(filename)

#南山医院模板
def func_d(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['D%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['E%d' % i].value = dict_info.get('goods_name')
				t_sheet['F%d' % i].value = dict_info.get('phone')
				t_sheet['G%d' % i].value = dict_info.get('time')
				t_sheet['H%d' % i].value = dict_info.get('order_sn')
				t_sheet['I%d' % i].value = dict_info.get('shop_price')
				break
	t_wb.save(filename)

#港大医院模板
def func_e(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = get_birth_date(dict_info.get('manid'))
				t_sheet['D%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['E%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['F%d' % i].value = dict_info.get('phone')
				t_sheet['G%d' % i].value = dict_info.get('manid')
				t_sheet['H%d' % i].value = dict_info.get('goods_name')
				t_sheet['I%d' % i].value = dict_info.get('time')
				t_sheet['J%d' % i].value = dict_info.get('shop_price')
				t_sheet['K%d' % i].value = '下午'
				break
	t_wb.save(filename)

#儿童医院模板	
def func_f(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('time')
				t_sheet['B%d' % i].value = dict_info.get('goods_name')
				t_sheet['C%d' % i].value = dict_info.get('shop_price')
				t_sheet['D%d' % i].value = dict_info.get('name')
				t_sheet['E%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['F%d' % i].value = dict_info.get('birthday')
				t_sheet['G%d' % i].value = dict_info.get('manid')
				t_sheet['H%d' % i].value = dict_info.get('phone')
				t_sheet['I%d' % i].value = dict_info.get('fa_name')
				t_sheet['J%d' % i].value = dict_info.get('receive_addr')
				break
	t_wb.save(filename)
	
#妇幼保健院模板	
def func_g(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['D%d' % i].value = dict_info.get('goods_name')
				t_sheet['E%d' % i].value = ''
				t_sheet['F%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['G%d' % i].value = dict_info.get('phone')
				t_sheet['H%d' % i].value = dict_info.get('manid')
				t_sheet['I%d' % i].value = dict_info.get('time')
				break
	t_wb.save(filename)

#乐检查模板
def func_h(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['D%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['E%d' % i].value = dict_info.get('manid')
				t_sheet['F%d' % i].value = dict_info.get('name')
				t_sheet['G%d' % i].value = dict_info.get('phone')
				t_sheet['H%d' % i].value = dict_info.get('goods_name')
				t_sheet['I%d' % i].value = ''
				t_sheet['J%d' % i].value = dict_info.get('time')
				t_sheet['K%d' % i].value = dict_info.get('shop_price')
				t_sheet['L%d' % i].value = dict_info.get('org_name')
				break
	t_wb.save(filename)

#成都市第三人民医院模板
def func_i(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = get_age(dict_info.get('manid'))
				t_sheet['D%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['E%d' % i].value = dict_info.get('phone')
				t_sheet['F%d' % i].value = dict_info.get('manid')
				t_sheet['G%d' % i].value = dict_info.get('time')
				t_sheet['H%d' % i].value = dict_info.get('goods_name')
				t_sheet['I%d' % i].value = dict_info.get('shop_price')
				break
	t_wb.save(filename)
	
#爱康国宾模板
def func_j(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = '无'
				t_sheet['B%d' % i].value = '无'
				t_sheet['C%d' % i].value = dict_info.get('goods_name')
				t_sheet['D%d' % i].value = dict_info.get('name')
				t_sheet['E%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['F%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['G%d' % i].value = get_birth_date(dict_info.get('manid'))
				t_sheet['H%d' % i].value = dict_info.get('manid')
				t_sheet['I%d' % i].value = dict_info.get('phone')
				t_sheet['J%d' % i].value = dict_info.get('org_name')
				t_sheet['K%d' % i].value = dict_info.get('time')
				break
	t_wb.save(filename)
	
#瑞慈体检中心模板
def func_k(list_info,filename):
	t_wb = load_workbook(filename)
	t_sheet = t_wb.get_sheet_by_name('Sheet1')
	
	for dict_info in list_info:
		i = 1
		while(True):
			if(t_sheet['A%d' % i].value):
				i += 1
				continue
			else:
				t_sheet['A%d' % i].value = dict_info.get('name')
				t_sheet['B%d' % i].value = "男" if dict_info.get('sex')=='0' else "女"
				t_sheet['C%d' % i].value = "已婚" if dict_info.get('marry') == '1' else "未婚"
				t_sheet['D%d' % i].value = dict_info.get('manid')
				t_sheet['E%d' % i].value = dict_info.get('phone')
				t_sheet['F%d' % i].value = ''
				t_sheet['G%d' % i].value = ''
				t_sheet['H%d' % i].value = ''
				t_sheet['I%d' % i].value = ''
				t_sheet['J%d' % i].value = dict_info.get('org_name')
				t_sheet['K%d' % i].value = dict_info.get('time')
				t_sheet['L%d' % i].value = ''
				break
	t_wb.save(filename)
	
#	
def export_appoint_info(info,org_info):
	city_name = ''
	for city in citys:
		if org_info.get('cid') == city.get('id'):
			city_name = city.get('cname')
	list_info = json.loads(info)	#json转dict		-- json.dumps(info)		#dict转json
	cur_date = time.strftime('%Y%m%d')	#当前日期
	totalfilename = '预约信息/%s/%s/all-%s.xlsx' % (city_name,org_info.get('oname'),org_info.get('oname'))
	print(org_info.get('muban'))
	print(totalfilename)
	func = mubans.get(org_info.get('muban'))
	eval(func)(list_info,totalfilename)
	
#身份证转换年龄
def get_age(manid):
	try:
		if checkIdcard(manid):
			year = int(manid[6:10])
			cur_year = int(time.strftime("%Y"))
			return (cur_year - year)
		else:
			return ''
	except Exception:
		return ''
	
#身份证转换出生日期
def get_birth_date(manid):
	try:
		if checkIdcard(manid):
			d = [0,0,0]
			d[0] = manid[6:10]
			d[1] = manid[10:12]
			d[2] = manid[12:14]
			return '/'.join(d)
		else:
			return ''
	except Exception:
		return ''

#身份证校验
#Errors=['验证通过!','身份证号码位数不对!','身份证号码出生日期超出范围或含有非法字符!','身份证号码校验错误!','身份证地区非法!']
def checkIdcard(idcard):
    Errors=['验证通过!','身份证号码位数不对!','身份证号码出生日期超出范围或含有非法字符!','身份证号码校验错误!','身份证地区非法!']
    area={"11":"北京","12":"天津","13":"河北","14":"山西","15":"内蒙古","21":"辽宁","22":"吉林","23":"黑龙江","31":"上海","32":"江苏","33":"浙江","34":"安徽","35":"福建","36":"江西","37":"山东","41":"河南","42":"湖北","43":"湖南","44":"广东","45":"广西","46":"海南","50":"重庆","51":"四川","52":"贵州","53":"云南","54":"西藏","61":"陕西","62":"甘肃","63":"青海","64":"宁夏","65":"新疆","71":"台湾","81":"香港","82":"澳门","91":"国外"}
    idcard=str(idcard)
    idcard=idcard.strip()
    idcard_list=list(idcard)

    #地区校验
    if(not area[(idcard)[0:2]]):
        print(Errors[4])
        return False
    #15位身份号码检测
    if(len(idcard)==15):
        if((int(idcard[6:8])+1900) % 4 == 0 or((int(idcard[6:8])+1900) % 100 == 0 and (int(idcard[6:8])+1900) % 4 == 0 )):
            erg=re.compile('[1-9][0-9]{5}[0-9]{2}((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|[1-2][0-9]))[0-9]{3}$')#//测试出生日期的合法性
        else:
            ereg=re.compile('[1-9][0-9]{5}[0-9]{2}((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|1[0-9]|2[0-8]))[0-9]{3}$')#//测试出生日期的合法性
        if(re.match(ereg,idcard)):
            print(Errors[0])
            return True
        else:
            print(Errors[2])
            return False
    #18位身份号码检测
    elif(len(idcard)==18):
        #出生日期的合法性检查
        #闰年月日:((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|[1-2][0-9]))
        #平年月日:((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|1[0-9]|2[0-8]))
        if(int(idcard[6:10]) % 4 == 0 or (int(idcard[6:10]) % 100 == 0 and int(idcard[6:10])%4 == 0 )):
            ereg=re.compile('[1-9][0-9]{5}19[0-9]{2}((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|[1-2][0-9]))[0-9]{3}[0-9Xx]$')#//闰年出生日期的合法性正则表达式
        else:
            ereg=re.compile('[1-9][0-9]{5}19[0-9]{2}((01|03|05|07|08|10|12)(0[1-9]|[1-2][0-9]|3[0-1])|(04|06|09|11)(0[1-9]|[1-2][0-9]|30)|02(0[1-9]|1[0-9]|2[0-8]))[0-9]{3}[0-9Xx]$')#//平年出生日期的合法性正则表达式
        #//测试出生日期的合法性
        if(re.match(ereg,idcard)):
            #//计算校验位
            S = (int(idcard_list[0]) + int(idcard_list[10])) * 7 + (int(idcard_list[1]) + int(idcard_list[11])) * 9 + (int(idcard_list[2]) + int(idcard_list[12])) * 10 + (int(idcard_list[3]) + int(idcard_list[13])) * 5 + (int(idcard_list[4]) + int(idcard_list[14])) * 8 + (int(idcard_list[5]) + int(idcard_list[15])) * 4 + (int(idcard_list[6]) + int(idcard_list[16])) * 2 + int(idcard_list[7]) * 1 + int(idcard_list[8]) * 6 + int(idcard_list[9]) * 3
            Y = S % 11
            M = "F"
            JYM = "10X98765432"
            M = JYM[Y]#判断校验位
            idcard_list[17]='X' if idcard_list[17] == 'x' else idcard_list[17] 	# add lj
            if(M == idcard_list[17]):#检测ID的校验位
                print(Errors[0])
                return True
            else:
                print(Errors[3])
                return False
        else:
            print(Errors[2])
            return False
    else:
        print(Errors[1])
        return False	

if __name__ == '__main__':
	if not os.path.exists('./param/provinces.d'):
		print('配置文件出错..')
	if not os.path.exists('./param/citys.d'):
		print('配置文件出错..')
	if not os.path.exists('./param/orgs.d'):
		print('配置文件出错..')
	if not os.path.exists('./param/mubans.d'):
		print('配置文件出错..')
		
	with open('./param/provinces.d','r',encoding='utf-8') as f:
		#print(f)
		provinces = json.loads(f.read())
		#print(provinces)
		
	with open('./param/citys.d','r',encoding='utf-8') as f:
		#print(f)
		citys = json.loads(f.read())
		#print(citys)
	
	with open('./param/orgs.d','r',encoding='utf-8') as f:
		#print(f.read())
		orgs = json.loads(f.read())

	MyFrame()
